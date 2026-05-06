#!/usr/bin/env python3
"""
mock_students.py

Convert a Learn survey "Summary Report" (CSV or XLSX) into a per-student
mock dataset suitable for testing the team formation algorithm.

Output columns:
    student_number, allocation_category, studyline, personality_type

Student count is driven by Question 2 (study programme): if 14 people picked
Civil Engineering, the script produces 14 students with that studyline.
Personalities from Question 3 are shuffled and paired across the students so
studyline and personality are not artificially correlated.

Usage:
    python mock_students.py path/to/summary.csv
    python mock_students.py path/to/summary.xlsx -o students.csv -c A --seed 42

Note: XLSX/XLSM support requires `openpyxl` (pip install openpyxl).
"""

import argparse
import csv
import random
import re
import sys
from pathlib import Path

# Question numbers as they appear in the "Q #" column of the summary CSV.
Q_STUDYLINE = "2"
Q_PERSONALITY = "3"


def _load_rows_csv(path):
    # utf-8-sig handles the BOM that Learn exports tend to ship with.
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            yield row


def _load_rows_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("Reading .xlsx requires openpyxl. Install it with: "
                 "pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active  # Learn exports keep the report on the first sheet
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    for raw in rows:
        if raw is None or all(v is None for v in raw):
            continue
        yield {h: v for h, v in zip(headers, raw)}


def load_rows(path):
    """Yield dict rows from either a CSV or an XLSX/XLSM file."""
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        return _load_rows_csv(path)
    if ext in (".xlsx", ".xlsm"):
        return _load_rows_xlsx(path)
    sys.exit(f"Unsupported file format: {ext} "
             "(supported: .csv, .xlsx, .xlsm — convert .xls/.xlsb to .xlsx first)")


def parse_summary(path):
    """Return (studyline_counts, personality_counts) as [(label, n), ...]."""
    studyline_counts = []
    personality_counts = []
    for row in load_rows(path):
        q = str(row.get("Q #") or "").strip()
        label = str(row.get("Answer") or "").strip()
        raw_n = row.get("# Responses") or 0
        try:
            n = int(raw_n)
        except (TypeError, ValueError):
            n = 0
        if n <= 0 or not label:
            continue
        if q == Q_STUDYLINE:
            studyline_counts.append((label, n))
        elif q == Q_PERSONALITY:
            personality_counts.append((label, n))
    return studyline_counts, personality_counts


def expand(counts):
    """Turn [(label, n), ...] into a flat list with each label repeated n times."""
    out = []
    for label, n in counts:
        out.extend([label] * n)
    return out


def detect_challenge(filename):
    """Pick 'Challenge X' out of the filename. Defaults to 'A' if not found."""
    m = re.search(r"Challenge[_\s\-]*([A-E])", filename, re.IGNORECASE)
    return m.group(1).upper() if m else "A"


def make_students(studylines, personalities, allocation_category,
                  start_id=0, seed=None):
    """Pair studylines with personalities and assign sequential student numbers."""
    rng = random.Random(seed)

    # The two lists ought to be the same length, but real exports get messy.
    # Pad with random picks (or truncate) so we always end up with one
    # personality per student. Mismatches are reported by the caller.
    n_sl = len(studylines)
    n_p = len(personalities)
    if n_p < n_sl and n_p > 0:
        personalities = personalities + [rng.choice(personalities)
                                         for _ in range(n_sl - n_p)]
    elif n_p > n_sl:
        personalities = personalities[:n_sl]

    # Shuffle so e.g. all the Biotech students don't end up with the same MBTI.
    rng.shuffle(personalities)

    students = []
    for i, (sl, p) in enumerate(zip(studylines, personalities)):
        sid = f"s{start_id + i:06d}"
        students.append({
            "student_number": sid,
            "allocation_category": allocation_category,
            "studyline": sl,
            "personality_type": p,
        })
    return students


def write_students(students, out_path):
    fieldnames = ["student_number", "allocation_category",
                  "studyline", "personality_type"]
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(students)


def main():
    ap = argparse.ArgumentParser(
        description="Expand a survey summary file (CSV or XLSX) into per-student mock records."
    )
    ap.add_argument("input", help="Path to the survey summary (.csv, .xlsx, or .xlsm)")
    ap.add_argument("-o", "--output",
                    help="Output CSV path (default: <input_stem>_students.csv)")
    ap.add_argument("-c", "--challenge",
                    help="Challenge letter (A-E). Auto-detected from filename if omitted.")
    ap.add_argument("--start-id", type=int, default=0,
                    help="First student-number suffix (default: 0)")
    ap.add_argument("--seed", type=int, default=42,
                    help="Random seed for reproducible shuffling (default: 42)")
    args = ap.parse_args()

    in_path = Path(args.input)
    if not in_path.exists():
        sys.exit(f"Input file not found: {in_path}")

    out_path = (Path(args.output) if args.output
                else in_path.with_name(f"{in_path.stem}_students.csv"))
    challenge = (args.challenge or detect_challenge(in_path.name)).upper()
    allocation_category = f"challenge {challenge}"

    studyline_counts, personality_counts = parse_summary(in_path)
    studylines = expand(studyline_counts)
    personalities = expand(personality_counts)

    n_sl = len(studylines)
    n_p = len(personalities)
    print(f"Allocation category    : {allocation_category}")
    print(f"Studyline responses    : {n_sl}")
    print(f"Personality responses  : {n_p}")
    if n_sl != n_p:
        print("  (mismatch — personalities will be padded/truncated to match)")

    students = make_students(studylines, personalities, allocation_category,
                             start_id=args.start_id, seed=args.seed)
    write_students(students, out_path)
    print(f"Wrote {len(students)} students -> {out_path}")


if __name__ == "__main__":
    main()
