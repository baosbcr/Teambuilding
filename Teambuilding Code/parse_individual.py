#!/usr/bin/env python3
"""
Parse Learn "Individual Attempts" exports into a per-student CSV.

Usage:
    python parse_individual.py file1.xlsx file2.xlsx -o students.csv
    python parse_individual.py oddly_named.xlsx -c overflow
"""

import argparse
import csv
import re
import sys
from pathlib import Path

Q_STUDENT_NUMBER = "1"
Q_STUDYLINE      = "2"
Q_PERSONALITY    = "3"

# Tried in order, first match wins. None as the result means read the letter from the capture group.
_CATEGORY_PATTERNS = [
    (re.compile(r"(challenge[\s_-]*)?overflow",        re.IGNORECASE), "overflow"),
    (re.compile(r"late[\s_-]*entr(y|ies)",             re.IGNORECASE), "late entry"),
    (re.compile(r"challenge[\s_-]*([A-D])(?![a-zA-Z])", re.IGNORECASE), None),
]


def detect_category(filename):
    stem = Path(filename).stem
    for pattern, category in _CATEGORY_PATTERNS:
        m = pattern.search(stem)
        if m:
            return category if category else f"challenge {m.group(1).upper()}"
    print(f"WARNING: could not detect category from '{filename}' — use -c to set it manually", file=sys.stderr)
    return None


def _rows_from_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        yield from csv.DictReader(f)


def _rows_from_xlsx(path):
    from openpyxl import load_workbook  # type: ignore[import-untyped]
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        raw_rows = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(raw_rows)]
        for raw in raw_rows:
            if raw is None or all(v is None for v in raw):
                continue
            yield {h: (str(v).strip() if v is not None else "") for h, v in zip(headers, raw)}
    finally:
        wb.close()


def load_rows(path):
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        return list(_rows_from_csv(path))
    if ext in (".xlsx", ".xlsm"):
        return list(_rows_from_xlsx(path))
    sys.exit(f"Unsupported format: {ext}  (supported: .csv, .xlsx, .xlsm)")


def split_into_blocks(rows):
    if not rows:
        return
    first_key = next(iter(rows[0]))
    current_name, current_block = None, []
    for row in rows:
        q_num     = (row.get("Q #")      or "").strip()
        first_val = (row.get(first_key)  or "").strip()
        if first_val and not q_num:
            if current_name is not None:
                yield current_name, current_block
            current_name, current_block = first_val, []
        else:
            current_block.append(row)
    if current_name is not None:
        yield current_name, current_block


def parse_block(name, rows):
    student_number       = None
    studylines_selected  = []
    personalities_selected = []

    for row in rows:
        q      = (row.get("Q #")    or "").strip()
        answer = (row.get("Answer") or "").strip()
        try:
            resp = int(row.get("# Responses") or 0)
        except (TypeError, ValueError):
            resp = 0

        if q == Q_STUDENT_NUMBER:
            raw = (row.get("Answer Match") or row.get("Answer") or "").strip()
            student_number = raw or None
        elif q == Q_STUDYLINE and resp == 1:
            studylines_selected.append(answer)
        elif q == Q_PERSONALITY and resp == 1:
            personalities_selected.append(answer)

    warnings = []
    if not student_number:
        warnings.append("missing student number")
    if not studylines_selected:
        warnings.append("no studyline selected")
    elif len(studylines_selected) > 1:
        warnings.append(f"multiple studylines {studylines_selected} -- using first")
    if not personalities_selected:
        warnings.append("no personality selected")
    elif len(personalities_selected) > 1:
        warnings.append(f"multiple personalities {personalities_selected} -- using first")

    label = student_number or name
    for w in warnings:
        print(f"WARNING [{label}]: {w}", file=sys.stderr)

    if not student_number and not studylines_selected and not personalities_selected:
        print(f"SKIPPED [{name}]: empty block", file=sys.stderr)
        return None

    return {
        "student_name":     name,
        "student_number":   student_number or name,
        "studyline":        studylines_selected[0]     if studylines_selected     else "UNKNOWN",
        "personality_type": personalities_selected[0]  if personalities_selected  else "UNKNOWN",
    }


def process_file(in_path, category):
    rows, students, skipped = load_rows(in_path), [], 0
    for name, block_rows in split_into_blocks(rows):
        record = parse_block(name, block_rows)
        if record is None:
            skipped += 1
        else:
            students.append(record)
    print(f"  {in_path.name}")
    print(f"    category : {category}")
    print(f"    students : {len(students)}" + (f"  (skipped {skipped})" if skipped else ""))
    return students


def load_all_surveys(survey_dir: Path) -> list[dict]:
    """Read every survey XLSX/CSV in survey_dir; return flat list of survey records."""
    in_paths = sorted(
        p for p in survey_dir.iterdir()
        if p.suffix.lower() in (".xlsx", ".xlsm", ".csv")
        and not p.name.startswith("~$")
    )
    if not in_paths:
        sys.exit(f"No survey files found in '{survey_dir}'")
    records: list[dict] = []
    for p in in_paths:
        cat = detect_category(p.name)
        if cat is None:
            sys.exit(
                f"Could not detect category for '{p.name}' - "
                "rename the file or check _CATEGORY_PATTERNS in parse_individual.py"
            )
        for s in process_file(p, cat):
            s["allocation_category"] = cat
            records.append(s)
    return records


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("inputs", nargs="+", help="One or more .xlsx/.csv files, or a single directory")
    ap.add_argument("-o", "--output",   help="Output CSV (default: auto-named)")
    ap.add_argument("-c", "--category", help="Override category when filename detection fails")
    args = ap.parse_args()

    # Resolve output path first so we can exclude it from the directory scan
    out_path = Path(args.output) if args.output else Path("students_combined.csv")

    # If a single directory is passed, expand it to all supported files inside
    if len(args.inputs) == 1 and Path(args.inputs[0]).is_dir():
        folder   = Path(args.inputs[0])
        in_paths = sorted(
            p for p in folder.iterdir()
            if p.suffix.lower() in (".xlsx", ".xlsm", ".csv")
            and p.resolve() != out_path.resolve()
            and not p.name.startswith("~$")  # skip Office lock files
        )
        if not in_paths:
            sys.exit(f"No .xlsx/.xlsm/.csv files found in '{folder}'")
    else:
        in_paths = [Path(p) for p in args.inputs]
        missing  = [p for p in in_paths if not p.exists()]
        if missing:
            sys.exit("File(s) not found:\n" + "\n".join(f"  {p}" for p in missing))
        # For a single explicit file with no -o, name the output next to it
        if not args.output and len(in_paths) == 1:
            out_path = in_paths[0].with_name(f"{in_paths[0].stem}_students.csv")

    if args.category and len(in_paths) > 1:
        sys.exit("-c only applies to a single file — run each file separately if categories differ")

    all_students = []
    print("Parsing files:")
    for in_path in in_paths:
        category = args.category or detect_category(in_path.name)
        if category is None:
            sys.exit(f"Could not determine category for '{in_path.name}' — re-run with -c")
        for s in process_file(in_path, category):
            s["allocation_category"] = category
            all_students.append(s)

    fieldnames = ["student_number", "student_name", "allocation_category",
                  "studyline", "personality_type"]
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(all_students)

    print(f"\nTotal : {len(all_students)} students -> {out_path}")


if __name__ == "__main__":
    main()
